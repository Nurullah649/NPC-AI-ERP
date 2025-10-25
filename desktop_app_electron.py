# -*- coding: utf-8 -*-
# Gerekli Kütüphaneler: selenium, requests, openpyxl, python-dotenv, thefuzz, python-docx, googletrans==3.0.0, langdetect, chardet
import sys
import os
import time
# import hashlib
# import requests  # Zaten vardır muhtemelen

# # Sadece Windows'ta çalışacak donanım ID'si için
# try:
#     import wmi
# except ImportError:
#     # Windows dışı bir sistemde (veya wmi yüklü değilse)
#     # test için sahte bir sınıf oluştur
#     class MockWMI:
#         def Win32_BaseBoard(self):
#             return [type('obj', (object,), {'SerialNumber': 'TEST-SERIAL-NUMBER'})]
#     wmi = type('wmi', (object,), {'WMI': MockWMI})
import json
import logging
import re
import threading
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, Any, List
from pathlib import Path
import openpyxl
import docx
import csv
import chardet
from dotenv import load_dotenv
from thefuzz import fuzz
import io
from openpyxl.styles import Font, Alignment

from googletrans import Translator
from langdetect import detect, LangDetectException


# src modüllerinin import edildiği varsayılıyor
# from src import sigma, netflex, tci, currency_converter, orkim

# --- ÖRNEK/BOŞ MODÜLLER (Eğer 'src' klasörü yoksa hata vermemesi için) ---
class MockAPI:
    def __init__(self, *args, **kwargs): pass

    def search_products(self, *args, **kwargs): return []

    def get_all_products(self, *args, **kwargs): return []  # for ITK

    def start_drivers(self, *args, **kwargs): pass

    def stop_drivers(self, *args, **kwargs): pass

    def reinit_driver(self, *args, **kwargs): pass

    def close_driver(self, *args, **kwargs): pass

    def get_token(self, *args, **kwargs): pass

    def update_credentials(self, *args, **kwargs): pass

    def get_products(self, *args, **kwargs): yield [];

    def get_all_product_prices(self, *args, **kwargs): return {}

    # YENİ: Mock stok sorgulama fonksiyonu
    def _get_stock_from_page(self, *args, **kwargs): return 99


class MockConverter:
    def get_parities(self, *args, **kwargs): return {"usd_eur": 0.9, "gbp_eur": 1.1}


class MockProduct:
    def __init__(self, name, code, cas):
        self.name = name
        self.code = code
        self.cas_number = cas
        self.variations = []


# Gerçek importlar yerine Mock sınıfları kullanılıyor (geçici)
try:
    from src import sigma, netflex, tci, currency_converter, orkim, itk
except ImportError:
    print("Uyarı: 'src' modülleri bulunamadı. Mock sınıflar kullanılıyor.", file=sys.stderr)
    sigma = type('sigma', (), {'SigmaAldrichAPI': MockAPI})
    netflex = type('netflex', (), {'NetflexAPI': MockAPI, 'AuthenticationError': Exception})
    tci = type('tci', (), {'TciScraper': MockAPI, 'Product': MockProduct})
    currency_converter = type('currency_converter', (), {'CurrencyConverter': MockConverter})
    orkim = type('orkim', (), {'OrkimScraper': MockAPI})
    itk = type('itk', (), {'ItkScraper': MockAPI})


# --- BİTİŞ: ÖRNEK/BOŞ MODÜLLER ---


# --- PAKETLEME İÇİN DOSYA YOLU FONKSİYONU ---
def get_resource_path(relative_path: str) -> str:
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# --- YENİ VE GÜVENLİ UYGULAMA VERİ YOLU TANIMI ---
def get_persistent_data_path() -> Path:
    if len(sys.argv) > 1:
        # Electron, veri yolunu bir argüman olarak gönderir.
        path_from_electron = Path(sys.argv[1])
        return path_from_electron
    # Eğer argüman gelmezse (örn: direkt python ile test ederken)
    elif sys.platform == "win32":
        # Windows için standart AppData klasörünü kullan
        return Path(os.getenv("APPDATA")) / "NPC-AI-ERP"
    else:
        # Diğer sistemler için (macOS, Linux) fallback
        return Path.home() / ".config" / "NPC-AI-ERP"


# --- GÜVENLİ VERİ YOLU ATAMALARI ---
# Artık tüm kullanıcı verileri (ayarlar, loglar, takvim) güncellemelerden etkilenmeyen
# bu güvenli klasöre kaydedilecek.
LOGS_AND_SETTINGS_DIR = get_persistent_data_path()
SETTINGS_FILE_PATH = LOGS_AND_SETTINGS_DIR / "settings.json"
CALENDAR_NOTES_FILE_PATH = LOGS_AND_SETTINGS_DIR / "calendar_notes.json"
NOTIFICATION_STATE_FILE = LOGS_AND_SETTINGS_DIR / "notification_state.json"

# .env dosyası hala paket içinden okunuyor, bu doğru.
dotenv_path = get_resource_path('.env')
load_dotenv(dotenv_path=dotenv_path)

# Bildirim zamanlayıcı için global değişkenler
notification_thread = None
notification_running = False

# ITK ürünleri için global önbellek
itk_product_cache = []
itk_cache_lock = threading.Lock()


#
# # --- YENİ: Lisans ve Donanım ID Fonksiyonları ---
#
# def get_hashed_device_id() -> str:
#     """
#     Anakart seri numarasını alır ve bunu SHA-256 ile hash'ler.
#     Bu, bilgisayar için benzersiz bir "parmak izi" oluşturur.
#     """
#     try:
#         c = wmi.WMI()
#         # Anakart seri numarasını al
#         serial = c.Win32_BaseBoard()[0].SerialNumber.strip()
#
#         # Seri numarasını hash'le
#         hashed_id = hashlib.sha256(serial.encode('utf-8')).hexdigest()
#         logging.info(f"Cihaz ID oluşturuldu. Hash: {hashed_id}")
#         return hashed_id
#     except Exception as e:
#         logging.error(f"Cihaz ID'si oluşturulamadı: {e}", exc_info=True)
#         # Hata durumunda veya (macOS/Linux'ta) sabit bir ID dön
#         # Bu, wmi'nin çalışmadığı durumlar için bir geri dönüştür
#         fallback_id = "error_or_non_windows_device_id"
#         return hashlib.sha256(fallback_id.encode('utf-8')).hexdigest()
#
#
# def check_license(license_key: str, device_id: str) -> bool:
#     """
#     Lisans anahtarını ve cihaz ID'sini Firebase Cloud Function'a gönderir.
#     """
#     # !!! BURAYI KENDİ CLOUD FUNCTION URL'NİZ İLE DEĞİŞTİRİN !!!
#     CLOUD_FUNCTION_URL = "https://YOUR-CLOUD-FUNCTION-URL.cloudfunctions.net/checkLicense"
#
#     if not license_key:
#         logging.warning("Lisans anahtarı girilmemiş.")
#         return False
#
#     try:
#         payload = {"licenseKey": license_key, "deviceId": device_id}
#         response = requests.post(CLOUD_FUNCTION_URL, json=payload, timeout=10) # Timeout artırıldı
#         if response.status_code == 200 and response.json().get("status") == "valid":
#             logging.info("Lisans doğrulandı (Sunucu yanıtı: Geçerli).")
#             return True
#         logging.warning(f"Lisans geçersiz (Sunucu yanıtı: {response.status_code} - {response.text}).")
#         return False
#     except requests.RequestException as e:
#         logging.error(f"Lisans sunucusuna bağlanırken hata: {e}", exc_info=True)
#         return False

# --- Ayarları Yükleme/Kaydetme Fonksiyonları ---
def load_settings() -> (Dict[str, Any], bool):  # YENİ: bool değeri döndürür (was_upgraded)
    """
    Kullanıcı ayarlarını yükler. Eğer ayar dosyası yoksa varsayılanları oluşturur.
    Mevcut ayar dosyasını, koddaki yeni varsayılan ayarlarla akıllıca günceller.
    """
    # Tüm ayarlar ve varsayılan değerleri burada tanımlanır.
    # Yeni bir ayar eklendiğinde, sadece bu listeye eklemek yeterlidir.
    default_settings = {
        "netflex_username": "", "netflex_password": "", "tci_coefficient": 1.4,
        "sigma_coefficient_us": 1.0, "sigma_coefficient_de": 1.0, "sigma_coefficient_gb": 1.0,
        "orkim_username": "", "orkim_password": "",
        "itk_username": "", "itk_password": "", "itk_coefficient": 1.0,
        # "license_key": ""  # YENİ SATIRI EKLEYİN
        # Örnek yeni ayar: "auto_check_for_updates": True
    }

    # Veri klasörünün mevcut olduğundan emin ol
    LOGS_AND_SETTINGS_DIR.mkdir(exist_ok=True)

    if not SETTINGS_FILE_PATH.exists():
        # Ayar dosyası yoksa, varsayılanları kaydet ve yükseltme olmadığını belirt.
        save_settings(default_settings)
        return default_settings, False

    try:
        with open(SETTINGS_FILE_PATH, 'r', encoding='utf-8') as f:
            user_settings = json.load(f)

        # Akıllı yükseltme: Varsayılanları temel al, kullanıcının ayarlarını üzerine yaz.
        # Bu, yeni eklenen varsayılan ayarların korunmasını sağlar.
        final_settings = default_settings.copy()
        final_settings.update(user_settings)

        # Yeni bir ayar eklenip eklenmediğini kontrol et (anahtar setleri farklı mı?)
        was_upgraded = set(final_settings.keys()) != set(user_settings.keys())

        # Eğer ayarlar yükseltildiyse (yeni anahtar eklendiyse), dosyayı güncelle.
        if was_upgraded:
            save_settings(final_settings)
            logging.info("Ayarlar yeni versiyona yükseltildi. Yeni ayarlar eklendi.")

        return final_settings, was_upgraded

    except (json.JSONDecodeError, IOError):
        # Hata durumunda varsayılanlara dön ve yükseltme olmadığını belirt.
        return default_settings, False


def save_settings(new_settings: Dict[str, Any]):
    try:
        for key in ['tci_coefficient', 'sigma_coefficient_us', 'sigma_coefficient_de', 'sigma_coefficient_gb',
                    'itk_coefficient']:
            if key in new_settings and new_settings.get(key):
                new_settings[key] = float(str(new_settings[key]).replace(',', '.'))
        LOGS_AND_SETTINGS_DIR.mkdir(exist_ok=True)
        with open(SETTINGS_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(new_settings, f, indent=4, ensure_ascii=False)
    except (IOError, TypeError, ValueError) as e:
        logging.error(f"Ayarlar kaydedilirken hata: {e}")


# --- TAKVİM VE BİLDİRİM FONKSİYONLARI ---
def load_calendar_notes() -> list:
    if not CALENDAR_NOTES_FILE_PATH.exists(): return []
    try:
        with open(CALENDAR_NOTES_FILE_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return []


def save_calendar_notes(notes: list):
    try:
        LOGS_AND_SETTINGS_DIR.mkdir(exist_ok=True)
        with open(CALENDAR_NOTES_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(notes, f, indent=4, ensure_ascii=False)
    except (IOError, TypeError) as e:
        logging.error(f"Takvim notları kaydedilirken hata: {e}")


# ... (Geri kalan tüm Python kodunuz burada hiçbir değişiklik olmadan devam ediyor) ...
# ...
# ... (Dosyanın sonuna kadar olan tüm kodları buraya yapıştırın)
# ...


def _mark_meeting_as_complete(note_date: str, meeting_id: str):
    try:
        notes = load_calendar_notes()
        for note in notes:
            if note.get("date") == note_date:
                for meeting in note.get("meetings", []):
                    if meeting.get("id") == meeting_id:
                        meeting["completed"] = True
                        save_calendar_notes(notes)
                        logging.info(f"Görüşme '{meeting_id}' tamamlandı olarak işaretlendi.")
                        send_to_frontend("calendar_notes_loaded", notes)
                        return
        logging.warning(f"Tamamlanacak görüşme bulunamadı: ID='{meeting_id}'")
    except Exception as e:
        logging.error(f"Görüşme tamamlanırken hata: {e}", exc_info=True)


def load_notification_state():
    if not NOTIFICATION_STATE_FILE.exists(): return {}
    try:
        with open(NOTIFICATION_STATE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {}


def save_notification_state(state):
    try:
        LOGS_AND_SETTINGS_DIR.mkdir(exist_ok=True)
        with open(NOTIFICATION_STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=4, ensure_ascii=False)
    except (IOError, TypeError) as e:
        logging.error(f"Bildirim durumu kaydedilirken hata: {e}")


def _perform_notification_check():
    try:
        logging.info("Periyodik bildirim kontrolü çalıştırılıyor...")
        notes = load_calendar_notes()
        state = load_notification_state()
        sent_ids = set(state.get("sent_ids", []))
        now = datetime.now()
        today = now.date()
        current_hour = now.hour

        thirty_days_ago = today - timedelta(days=30)
        sent_ids = {nid for nid in sent_ids if len(nid.split('_')) > 1 and datetime.strptime(nid.split('_')[1],
                                                                                             '%Y-%m-%d').date() >= thirty_days_ago}

        for note in notes:
            for meeting in note.get("meetings", []):
                meeting_type = meeting.get("type", "toplantı")  # Eski veriler için varsayılan "toplantı"
                frequency = meeting.get("notificationFrequency")
                meeting_date_str = meeting.get("nextMeetingDate")

                if meeting.get("completed") or not meeting_date_str or frequency == "none":
                    continue

                try:
                    meeting_date = datetime.strptime(meeting_date_str, "%Y-%m-%d").date()
                    should_notify_today = False

                    # Yeni "Görüşme" tipi için mantık
                    if meeting_type == 'görüşme' and frequency.startswith('for_'):
                        try:
                            parts = frequency.split('_')
                            if len(parts) == 3:
                                duration_val = int(parts[1])
                                duration_unit = parts[2]
                                delta = timedelta(days=0)
                                if duration_unit == 'day' or duration_unit == 'days':
                                    delta = timedelta(days=duration_val)
                                elif duration_unit == 'week' or duration_unit == 'weeks':
                                    delta = timedelta(weeks=duration_val)

                                start_date = meeting_date
                                end_date = meeting_date + delta - timedelta(days=1)
                                if start_date <= today <= end_date:
                                    should_notify_today = True
                        except (ValueError, IndexError):
                            logging.warning(f"Geçersiz 'görüşme' sıklık formatı: {frequency}")

                    # "Toplantı" tipi (ve eski veriler) için mantık
                    else:
                        notification_start_date = None
                        if frequency == "on_day":
                            notification_start_date = meeting_date
                        elif frequency == "1_day_before":
                            notification_start_date = meeting_date - timedelta(days=1)
                        elif frequency == "1_week_before":
                            notification_start_date = meeting_date - timedelta(days=7)

                        if notification_start_date and (notification_start_date <= today <= meeting_date):
                            should_notify_today = True

                    if should_notify_today:
                        daily_frequency = meeting.get("notificationDailyFrequency", "once")
                        notification_hours = []
                        if daily_frequency == "once":
                            notification_hours = [9]
                        elif daily_frequency == "twice":
                            notification_hours = [9, 17]
                        elif daily_frequency == "thrice":
                            notification_hours = [9, 13, 17]
                        elif daily_frequency == "five_times":
                            notification_hours = [9, 11, 13, 15, 17]
                        elif daily_frequency == "ten_times":
                            notification_hours = list(range(9, 19))
                        elif daily_frequency == "hourly":
                            notification_hours = list(range(9, 18))

                        if current_hour in notification_hours:
                            notif_id = f"{meeting.get('id')}_{today.strftime('%Y-%m-%d')}_{current_hour}"
                            if notif_id not in sent_ids:
                                company_name = meeting.get('companyName', meeting.get('personName', 'Bilinmeyen'))
                                notif_title = f"{meeting_type.capitalize()} Hatırlatması: {company_name}"
                                notif_subtitle = f"Tarih: {meeting_date.strftime('%d.%m.%Y')}"

                                logging.info(
                                    f"Bildirim tetikleniyor: {company_name} - Tip: {meeting_type} - Sıklık: {daily_frequency} - Saat: {current_hour}")
                                send_to_frontend("show_notification", {
                                    "title": notif_title,
                                    "subtitle": notif_subtitle,
                                    "body": meeting.get("meetingNotes", "Not eklenmemiş."),
                                    "noteDate": note.get("date"),
                                    "meetingId": meeting.get("id")
                                })
                                sent_ids.add(notif_id)

                except (ValueError, TypeError) as e:
                    logging.warning(f"Etkinlik işlenemedi. Veri: {meeting}. Hata: {e}")

        state["sent_ids"] = list(sent_ids)
        save_notification_state(state)
    except Exception as e:
        logging.error(f"Bildirim kontrolü sırasında hata: {e}", exc_info=True)


def check_and_send_notifications():
    global notification_running
    while notification_running:
        _perform_notification_check()
        for _ in range(3600):
            if not notification_running: break
            time.sleep(1)


def start_notification_scheduler():
    global notification_thread, notification_running
    if notification_running: return
    notification_running = True
    notification_thread = threading.Thread(target=check_and_send_notifications, daemon=True,
                                           name="Notification-Scheduler")
    notification_thread.start()
    logging.info("Bildirim zamanlayıcı başlatıldı.")


def stop_notification_scheduler():
    global notification_running
    if notification_running:
        notification_running = False
        logging.info("Bildirim zamanlayıcı durduruluyor...")
        if notification_thread and notification_thread.is_alive():
            notification_thread.join(2.0)
        logging.info("Bildirim zamanlayıcı durduruldu.")


# --- Loglama Yapılandırması ---
def setup_logging():
    for handler in logging.root.handlers[:]: logging.root.removeHandler(handler)
    formatter = logging.Formatter('%(asctime)s - [%(levelname)s] - (%(threadName)s) - %(message)s')
    log_dir = LOGS_AND_SETTINGS_DIR
    log_dir.mkdir(exist_ok=True)
    dev_log_file = log_dir / "developer.log"
    dev_handler = logging.FileHandler(dev_log_file, encoding='utf-8')
    dev_handler.setFormatter(formatter)
    dev_handler.setLevel(logging.INFO)
    admin_log_file = log_dir / "admin_activity.log"
    admin_handler = logging.FileHandler(admin_log_file, encoding='utf-8')
    admin_formatter = logging.Formatter('%(asctime)s - %(message)s')
    admin_handler.setFormatter(admin_formatter)
    admin_logger = logging.getLogger("admin")
    admin_logger.addHandler(admin_handler)
    admin_logger.setLevel(logging.INFO)
    admin_logger.propagate = False

    # Konsol çıktısını stderr'e yönlendiriyoruz
    console_handler = logging.StreamHandler(sys.stderr)
    console_handler.setFormatter(formatter)
    console_handler.setLevel(logging.INFO)

    logging.basicConfig(level=logging.INFO, handlers=[dev_handler, console_handler])
    for logger_name in ["urllib3", "selenium", "googletrans"]: logging.getLogger(logger_name).setLevel(logging.WARNING)
    return admin_logger


admin_logger = setup_logging()


# --- JSON Mesajlaşma (DÜZELTİLDİ) ---
def send_to_frontend(message_type: str, data: Any, context: Dict = None):
    try:
        message_obj = {"type": message_type, "data": data}
        if context: message_obj["context"] = context

        # Standart print() yerine doğrudan stdout buffer'ına UTF-8 byte yazıyoruz.
        json_string = json.dumps(message_obj, ensure_ascii=False) + '\n'
        sys.stdout.buffer.write(json_string.encode('utf-8'))
        sys.stdout.flush()
    except (TypeError, OSError, BrokenPipeError) as e:
        # Kapatma sırasında BrokenPipeError normaldir, loglamaya gerek yok.
        if not isinstance(e, BrokenPipeError):
            logging.error(f"Frontend'e mesaj gönderilemedi: {e}")


# --- Diğer Yardımcı Fonksiyonlar (Excel, Dosya Okuma, vb.) ---
def export_meetings_to_excel(data: Dict[str, Any]):
    notes = data.get("notes", [])
    start_date_str = data.get("startDate")
    end_date_str = data.get("endDate")

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return {"status": "error", "message": "Geçersiz tarih formatı."}

    meetings_to_export = []
    all_meetings = []
    for note in notes:
        all_meetings.extend(note.get("meetings", []))

    unique_meetings = {m['id']: m for m in all_meetings}.values()

    for meeting in unique_meetings:
        try:
            meeting_date = None
            if meeting.get("nextMeetingDate"):
                meeting_date = datetime.strptime(meeting.get("nextMeetingDate"), "%Y-%m-%d").date()

            if meeting_date and start_date <= meeting_date <= end_date:
                meeting['actual_meeting_date'] = meeting_date
                # Kayıt tarihini bulmak için notları tekrar gez
                for note in notes:
                    if any(m['id'] == meeting['id'] for m in note.get('meetings', [])):
                        meeting['note_date'] = datetime.strptime(note.get("date"), "%Y-%m-%d").date()
                        break
                else:  # Eğer bulunamazsa
                    meeting['note_date'] = meeting_date  # Yaklaşık bir tarih ata

                meetings_to_export.append(meeting)
        except (ValueError, TypeError):
            continue

    if not meetings_to_export:
        return {"status": "info", "message": "Belirtilen tarih aralığında dışa aktarılacak etkinlik bulunamadı."}

    meetings_to_export.sort(key=lambda m: m['actual_meeting_date'])

    desktop_path = Path.home() / "Desktop"
    desktop_path.mkdir(exist_ok=True)
    filename = f"Etkinlik_Raporu_{start_date_str}_-_{end_date_str}.xlsx"
    filepath = desktop_path / filename

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Etkinlik Listesi"

        headers = ["FİRMA ADI", "YETKİLİSİ", "DEPARTMANI", "MAİL ADRESİ", "TELEFON", "ETKİNLİK TİPİ", "KAYIT TARİHİ",
                   "ETKİNLİK TARİHİ",
                   "AÇIKLAMA"]
        sheet.append(headers)

        for cell in sheet["1:1"]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for meeting in meetings_to_export:
            formatted_meeting_date = meeting['actual_meeting_date'].strftime('%d.%m.%Y')
            formatted_note_date = meeting['note_date'].strftime('%d.%m.%Y') if 'note_date' in meeting else 'N/A'

            row = [
                meeting.get("companyName", ""),
                meeting.get("authorizedPerson", ""),
                meeting.get("department", ""),
                meeting.get("email", ""),
                meeting.get("phone", ""),
                meeting.get("type", "Bilinmiyor").capitalize(),
                formatted_note_date,
                formatted_meeting_date,
                meeting.get("meetingNotes", "")
            ]
            # Tüm string verilerin str() ile çevrildiğinden emin oluyoruz
            sheet.append([str(item) for item in row])

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(filepath)
        logging.info(f"Etkinlik listesi Excel dosyası oluşturuldu: {filepath}")
        return {"status": "success", "path": str(filepath)}
    except Exception as e:
        logging.error(f"Etkinlik Excel'i oluşturulurken hata: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}


def export_to_excel(data: Dict[str, Any]):
    customer_name = data.get("customerName", "Bilinmeyen_Musteri")
    products = data.get("products", [])
    safe_customer_name = re.sub(r'[\/*?:"<>|]', "", customer_name)
    desktop_path = Path.home() / "Desktop"
    desktop_path.mkdir(exist_ok=True)
    filename = f"{safe_customer_name}_urun_listesi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = desktop_path / filename
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Ürün Listesi"
        # --- YENİ BAŞLIK EKLEME ---
        headers = ["Kaynak", "Ürün Adı", "Marka", "Ürün Kodu", "Fiyat", "Para Birimi", "KDV", "Birim", "Stok Durumu"]
        # --- BAŞLIK GÜNCELLENDİ ---
        sheet.append(headers)
        for cell in sheet["1:1"]: cell.font = openpyxl.styles.Font(bold=True)

        for product in products:
            price_str_from_product = str(product.get("price_str", "N/A"))

            # KDV oranını ve temiz fiyat metnini ayıkla
            kdv_str = "Yok"  # Default KDV
            clean_price_str = price_str_from_product  # KDV'siz fiyat metni
            kdv_match = re.search(r'\+\s*%(\d+)\s*KDV', price_str_from_product, re.IGNORECASE)
            if kdv_match:
                kdv_str = f"%{kdv_match.group(1)}"
                # KDV kısmını fiyattan temizle
                clean_price_str = re.sub(r'\s*\+\s*%(\d+)\s*KDV.*', '', price_str_from_product,
                                         flags=re.IGNORECASE).strip()

            # Para birimi sembolünü temiz fiyattan belirle
            currency_symbol = ""
            price_str_lower = clean_price_str.lower()
            if '€' in price_str_lower or 'eur' in price_str_lower:
                currency_symbol = "€"
            elif '$' in price_str_lower or 'usd' in price_str_lower:
                currency_symbol = "$"
            elif '£' in price_str_lower or 'gbp' in price_str_lower:
                currency_symbol = "£"
            elif '₺' in price_str_lower or 'try' in price_str_lower or 'tl' in price_str_lower:
                currency_symbol = "₺"

            # Sayısal fiyatı (price_numeric) al, yoksa 0 ata
            price_val = product.get("price_numeric")
            excel_price_value = 0  # Default to 0 if no numeric price
            if isinstance(price_val, (int, float)):
                excel_price_value = price_val
            # Eğer sayısal fiyat yoksa ama temiz fiyattan parse edilebiliyorsa dene
            elif clean_price_str not in ["N/A", "Teklif İsteyiniz", ""]:
                try:
                    # Para birimi ve diğer karakterleri temizle
                    numeric_part = re.sub(r'[^\d,.]', '', clean_price_str).strip()
                    # Virgül ve nokta dönüşümü
                    if ',' in numeric_part and '.' in numeric_part:
                        if numeric_part.rfind(',') > numeric_part.rfind('.'):  # 1.234,56 formatı
                            numeric_part = numeric_part.replace('.', '').replace(',', '.')
                        else:  # 1,234.56 formatı
                            numeric_part = numeric_part.replace(',', '')
                    else:  # Sadece virgül varsa ondalık ayırıcıdır
                        numeric_part = numeric_part.replace(',', '.')

                    parsed_price = float(numeric_part)
                    if not isinstance(price_val, (int, float)):  # Sadece product.price_numeric yoksa bunu kullan
                        excel_price_value = parsed_price
                except ValueError:
                    pass  # Parse edilemezse 0 olarak kalır

            row = [
                product.get("source", "N/A"),
                product.get("product_name", "N/A"),
                product.get("brand", product.get("source", "N/A")),
                product.get("product_code", "N/A"),
                excel_price_value,
                currency_symbol,
                kdv_str,  # YENİ KDV SÜTUNU VERİSİ
                product.get("unit", "Adet"),
                product.get("cheapest_netflex_stock", "N/A")
            ]
            sheet.append(row)

        # Fiyat Sütunu Formatlama (E sütunu)
        price_column = sheet['E']
        for cell in price_column[1:]:  # Başlık hariç
            cell.number_format = '#,##0.00'

        # Sütun genişliklerini ayarla
        for col in sheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            try:
                # Sayısal sütunlar için genişlik hesaplamasını atla veya farklı yap
                if column_letter == 'E':  # Fiyat sütunuysa
                    max_length = 15  # Sabit bir genişlik ayarla
                else:
                    max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
            except (ValueError, TypeError):
                pass  # Boş sütun veya başka bir hata durumunda

            # Genişliği ayarla
            adjusted_width = max_length + 2
            if column_letter == 'E':
                adjusted_width = max_length  # Fiyat için sabit genişlik
            elif column_letter == 'G':  # KDV sütunu daha dar olabilir
                adjusted_width = 8

            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(filepath)
        logging.info(f"Excel dosyası oluşturuldu: {filepath}")
        admin_logger.info(f"Müşteri Ataması ve Rapor: Müşteri='{customer_name}', Atanan Ürün Sayısı={len(products)}")
        for product in products:
            p_name = re.sub(re.compile('<.*?>'), '', product.get("product_name", "N/A"))
            admin_logger.info(
                f"  -> Atanan Ürün: Ad='{p_name}', Kod='{product.get('product_code', 'N/A')}', Fiyat='{product.get('price_str', 'N/A')}'")
        return {"status": "success", "path": str(filepath)}
    except Exception as e:
        logging.error(f"Excel hatası: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}


translator = Translator()


def _translate_if_turkish(term: str) -> str:
    if not term: return term
    try:
        if detect(term) == 'tr':
            translated = translator.translate(term, src='tr', dest='en')
            if translated and translated.text:
                logging.info(f"Otomatik Çeviri: '{term}' -> '{translated.text}'")
                return translated.text
    except (LangDetectException, Exception):
        pass
    return term


def _clean_term(term):
    if not isinstance(term, str): return ""
    return re.sub(r'\s*$$[^)]*$$', '', term).strip()


def process_raw_data(data: List[List[str]]) -> List[str]:
    if not data: return []
    first_row_idx = next(
        (i for i, row in enumerate(data) if any(str(cell).strip() for cell in row if cell is not None)), -1)
    if first_row_idx == -1: return []
    relevant_data = data[first_row_idx:]
    potential_header = [str(cell).strip().lower() if cell is not None else '' for cell in relevant_data[0]]
    keywords = ['malzeme', 'ad', 'ürün', 'sarflar', 'proforma', 'açıklama', 'description', 'item', 'name',
                'stock keeping unit']
    target_col_idx = -1
    for i, header_cell in enumerate(potential_header):
        if any(keyword in header_cell for keyword in keywords):
            target_col_idx = i
            break
    body = relevant_data[1:] if target_col_idx != -1 else relevant_data
    if target_col_idx == -1: target_col_idx = next((i for i, cell in enumerate(potential_header) if cell), 0)
    search_terms = [str(row[target_col_idx]).strip() for row in body if
                    len(row) > target_col_idx and row[target_col_idx] and str(row[target_col_idx]).strip()]
    return search_terms


def read_excel_terms(file_path: str) -> List[str]:
    try:
        return process_raw_data(list(openpyxl.load_workbook(file_path, data_only=True).active.values))
    except Exception as e:
        logging.error(f"Excel okuma hatası: {e}", exc_info=True)
        return []


def read_docx_terms(file_path: str) -> List[str]:
    try:
        doc = docx.Document(file_path)
        return [term for table in doc.tables for term in
                process_raw_data([[cell.text for cell in row.cells] for row in table.rows])]
    except Exception as e:
        logging.error(f"Word okuma hatası: {e}", exc_info=True)
        return []


def read_csv_terms(file_path: str) -> List[str]:
    try:
        with open(file_path, 'rb') as f_raw:
            encoding = chardet.detect(f_raw.read())['encoding'] or 'utf-8'
        with open(file_path, 'r', encoding=encoding, newline='', errors='replace') as f:
            try:
                dialect = csv.Sniffer().sniff(f.read(2048))
            except csv.Error:
                dialect = 'excel'
            f.seek(0)
            return process_raw_data(list(csv.reader(f, dialect)))
    except Exception as e:
        logging.error(f"CSV okuma hatası: {e}", exc_info=True)
        return []


def get_search_terms_from_file(file_path):
    ext_map = {'.xlsx': read_excel_terms, '.csv': read_csv_terms, '.docx': read_docx_terms}
    file_ext = os.path.splitext(file_path)[1].lower()
    if file_ext not in ext_map: return []
    raw_terms = ext_map[file_ext](file_path)
    processed_terms = {_translate_if_turkish(_clean_term(term)) for term in raw_terms if _clean_term(term)}
    return [term for term in processed_terms if len(term) > 2]


# --- YENİ: Orkim Stok Sorgulama Görevi ---
def _get_orkim_stock_task(orkim_api_instance, product_url: str):
    """
    Ayrı bir thread'de Orkim'den detaylı stok sorgulaması yapar.
    orkim_api._get_stock_from_page yavaş bir fonksiyondur (sepete ekler, vb.)
    """
    try:
        if not orkim_api_instance:
            logging.warning("Orkim API başlatılmamışken stok sorgusu istendi.")
            send_to_frontend("orkim_stock_result", {"url": product_url, "stock": "Hata"})
            return

        logging.info(f"Orkim detaylı stok sorgusu başlatıldı: {product_url}")
        # Bu, orkim.py'deki yavaş, sepete ekleyen fonksiyondur
        stock_quantity = orkim_api_instance._get_stock_from_page(product_url)
        logging.info(f"Orkim detaylı stok sorgusu sonucu: {stock_quantity}")

        # Sonucu URL ile birlikte geri gönder (page.tsx'in doğru ürünü bulması için)
        send_to_frontend("orkim_stock_result", {"url": product_url, "stock": stock_quantity})

    except Exception as e:
        logging.error(f"Orkim stok sorgulama thread'inde hata ({product_url}): {e}", exc_info=True)
        send_to_frontend("orkim_stock_result", {"url": product_url, "stock": "Hata"})


# --- Karşılaştırma Motoru ---
class ComparisonEngine:
    def __init__(self, sigma_api: sigma.SigmaAldrichAPI, netflex_api: netflex.NetflexAPI, tci_api: tci.TciScraper,
                 orkim_api: orkim.OrkimScraper, itk_api: itk.ItkScraper,
                 initial_settings: Dict[str, Any], max_workers=10):
        self.sigma_api, self.netflex_api, self.tci_api, self.orkim_api, self.itk_api = sigma_api, netflex_api, tci_api, orkim_api, itk_api
        self.currency_converter = currency_converter.CurrencyConverter()
        self.max_workers = max_workers
        self.search_cancelled = threading.Event()
        self.batch_search_cancelled = threading.Event()
        self.settings = initial_settings

    def initialize_drivers(self):
        logging.info("Ağır servisler (Selenium sürücüleri) başlatılıyor...")
        start_time = time.monotonic()
        try:
            # Geliştirme modu takılma sorununu çözmek için sıralı başlatma
            logging.info("initialize_drivers BAŞLADI (Sıralı)")
            logging.info("Sigma sürücüleri başlatılıyor...")
            self.sigma_api.start_drivers()  # Bu kendi içinde thread kullanabilir, sorun olmaz
            logging.info("Sigma sürücüleri tamamlandı.")

            logging.info("TCI sürücüsü başlatılıyor...")
            self.tci_api.reinit_driver()
            logging.info("TCI sürücüsü tamamlandı.")

            # Orijinal paralel başlatma kodu (şimdilik yorumda)
            # with ThreadPoolExecutor(max_workers=2, thread_name_prefix="Heavy-Driver-Starter") as executor:
            #     f_sigma = executor.submit(self.sigma_api.start_drivers)
            #     f_tci = executor.submit(self.tci_api.reinit_driver)
            #     f_sigma.result()
            #     f_tci.result()
            logging.info(f"Tüm Selenium sürücüleri {time.monotonic() - start_time:.2f}s içinde başlatıldı (Sıralı).")
        except Exception as e:
            logging.critical(f"Selenium sürücüleri başlatılamadı: {e}", exc_info=True)
            raise e  # Hatanın yukarıya iletilip programın hata vermesini sağla

    # DEĞİŞİKLİK: search_data parametresi eklendi
    def _process_single_sigma_product_and_send(self, raw_sigma_product: Dict[str, Any], context: Dict,
                                               search_data: dict):
        # BU FONKSİYON İÇİNDEKİ "exact" MANTIĞI KULLANICININ İSTEĞİ ÜZERİNE DEĞİŞTİRİLMEDİ
        try:
            if self.search_cancelled.is_set(): return False
            s_num, s_brand, s_key, s_mids = raw_sigma_product.get('product_number'), raw_sigma_product.get(
                'brand'), raw_sigma_product.get('product_key'), raw_sigma_product.get('material_ids', [])

            # Sigma varyasyonlarını al
            sigma_variations_data = self.sigma_api.get_all_product_prices(s_num, s_brand, s_key.replace('.', ''),
                                                                          s_mids,
                                                                          self.search_cancelled)
            if self.search_cancelled.is_set(): return False

            # Netflex eşleşmeleri için terimleri topla
            netflex_terms = {s_num.replace('.', '')} if s_num else set()
            for country_vars in sigma_variations_data.values():
                # GÜNCELLEME: Varyasyonların liste olup olmadığını kontrol et
                if isinstance(country_vars, list):
                    for var in country_vars:
                        if isinstance(var, dict) and (mat_num := var.get('material_number')):
                            netflex_terms.add(mat_num.replace('.', ''))

            # Netflex araması yap ve sonuçları önbelleğe al
            netflex_cache = {}
            for term in netflex_terms:
                if self.search_cancelled.is_set(): return False
                # Dikkat: Netflex API'sinden AuthenticationError gelebilir, bunu yakalamalıyız
                try:
                    results = self.netflex_api.search_products(term, self.search_cancelled)
                    if results:
                        for r in results:
                            if r_code := r.get('product_code'): netflex_cache[r_code] = r
                except netflex.AuthenticationError:
                    logging.error(f"Netflex kimlik doğrulaması başarısız oldu (Sigma ürünü işlenirken). Ürün: {s_num}")
                    # Bu durumda Netflex aramasına devam etme, sadece Sigma sonuçlarıyla ilerle
                    break  # Netflex terim döngüsünden çık
                except Exception as e:
                    logging.error(f"Netflex araması sırasında beklenmedik hata (Sigma ürünü işlenirken {term}): {e}")

            if self.search_cancelled.is_set(): return False

            # Nihai ürünü oluştur
            final_product = self._build_final_sigma_product(raw_sigma_product, netflex_cache,
                                                            {s_num: sigma_variations_data},
                                                            self.settings)
            if final_product:
                # --- KULLANICININ İSTEDİĞİ ESNEK "exact" FİLTRELEME ---
                search_term = search_data.get("searchTerm", "").lower()
                search_logic = search_data.get("searchLogic", "exact")
                match_found = False

                if search_logic == "exact":
                    # --- KULLANICININ ORİJİNAL MANTIĞI KORUNDU ---
                    product_number_lower = final_product.get('product_number', '').lower()
                    product_name_lower = final_product.get('product_name', '').lower()
                    cas_number_lower = final_product.get('cas_number', '').lower()

                    if (search_term == product_name_lower or search_term == cas_number_lower):
                        match_found = True
                    elif (search_term in product_number_lower or product_number_lower in search_term):
                        match_found = True
                    elif (sigma_vars := final_product.get('sigma_variations')):
                        for country_vars in sigma_vars.values():
                            if isinstance(country_vars, list):
                                for var in country_vars:
                                    if isinstance(var, dict) and search_term == var.get('material_number', '').lower():
                                        match_found = True
                                        break
                            if match_found: break
                    elif (netflex_matches := final_product.get('netflex_matches')):
                        for match in netflex_matches:
                            if isinstance(match, dict) and search_term == match.get('product_code', '').lower():
                                match_found = True
                                break
                    # --- KULLANICININ ORİJİNAL MANTIĞI SONU ---
                else:  # similar
                    match_found = True

                if match_found:
                    send_to_frontend("product_found", {"product": final_product}, context=context)
                    return True  # Eşleşme bulundu ve gönderildi
                else:
                    logging.debug(f"Sigma ürünü '{s_num}' esnek exact filtreyi geçemedi ('{search_term}').")
                    return False  # Eşleşme bulunamadı

        except Exception as e:
            logging.error(f"Tekil Sigma ürünü ({raw_sigma_product.get('product_number')}) işlenirken hata: {e}",
                          exc_info=True)
        return False  # Hata durumunda veya eşleşme yoksa False dön

    def _build_final_sigma_product(self, sigma_product: Dict, netflex_cache: Dict, all_sigma_variations: Dict,
                                   settings: Dict) -> Dict or None:
        s_name, s_num, s_brand, cas = sigma_product.get('product_name_sigma'), sigma_product.get(
            'product_number'), sigma_product.get('brand'), sigma_product.get('cas_number', 'N/A')

        if not s_num: return None

        parities = self.currency_converter.get_parities()
        if "error" in parities: logging.error("Pariteler alınamadı.")

        sigma_variations = all_sigma_variations.get(s_num, {})
        all_price_options = []

        valid_sigma_variations_exist = False
        for country_code, variations in sigma_variations.items():
            coefficient = settings.get(f"sigma_coefficient_{country_code}", 1.0)
            if isinstance(variations, list):
                for var in variations:
                    if isinstance(var, dict) and 'error' not in var:
                        price_eur, original_price, currency = None, var.get('price'), var.get('currency', '').upper()
                        if original_price is not None:
                            try:
                                base_price_eur = None
                                if currency == 'USD' and parities.get('usd_eur'):
                                    base_price_eur = original_price * parities['usd_eur']
                                elif currency == 'GBP' and parities.get('gbp_eur'):
                                    base_price_eur = original_price * parities['gbp_eur']
                                elif currency == 'EUR':
                                    base_price_eur = original_price
                                if base_price_eur is not None:
                                    price_eur = base_price_eur * coefficient
                                    if mat_num := var.get('material_number'):
                                        all_price_options.append(
                                            {'price': price_eur, 'code': mat_num,
                                             'source': f"Sigma ({country_code.upper()})"})
                                        valid_sigma_variations_exist = True
                            except Exception as e:
                                logging.warning(
                                    f"Sigma fiyat hesaplamada hata ({s_num} - {var.get('material_number')}): {e}")

        netflex_matches = []
        sigma_mat_nums = set()
        for vars_list in sigma_variations.values():
            if isinstance(vars_list, list):
                for var in vars_list:
                    if isinstance(var, dict) and (mat_num := var.get('material_number')):
                        sigma_mat_nums.add(mat_num)

        if not sigma_mat_nums and s_num:
            sigma_mat_nums.add(s_num)

        for mat_num in sigma_mat_nums:
            clean_mat_num = mat_num.replace('.', '')
            if clean_mat_num in netflex_cache:
                match = netflex_cache[clean_mat_num]
                match['product_name'] = match.get('product_name') or s_name
                match['cas_number'] = cas
                netflex_matches.append(match)
                if price := match.get('price_numeric'):
                    all_price_options.append(
                        {'price': price, 'code': match.get('product_code'), 'source': 'Netflex'})

        cheapest_option = min(all_price_options, key=lambda x: x['price']) if all_price_options else {}

        final_product = {
            "source": "Sigma",
            "product_name": s_name or "N/A",
            "product_number": s_num,
            "cas_number": cas or "N/A",
            "brand": f"Sigma ({s_brand or 'N/A'})",
            "sigma_variations": sigma_variations if valid_sigma_variations_exist else {},
            "netflex_matches": netflex_matches,
            "cheapest_eur_price_str": cheapest_option.get('price') and f"{cheapest_option['price']:,.2f}€".replace(",",
                                                                                                                   "X").replace(
                ".", ",").replace("X", ".") or "N/A",
            "cheapest_material_number": cheapest_option.get('code', s_num),
            "cheapest_source_country": cheapest_option.get('source', "Netflex" if netflex_matches else "Sigma")
        }

        cheapest_code = final_product["cheapest_material_number"]
        cheapest_netflex_match = next((m for m in netflex_matches if m.get('product_code') == cheapest_code), None)
        final_product["cheapest_netflex_stock"] = cheapest_netflex_match.get('stock',
                                                                             'N/A') if cheapest_netflex_match else 'N/A'

        if not valid_sigma_variations_exist and netflex_matches:
            final_product["source"] = "Netflex (Sigma eşleşmesi)"

        if not all_price_options:
            final_product["cheapest_eur_price_str"] = "Fiyat Yok"
            final_product["cheapest_material_number"] = s_num
            final_product["cheapest_source_country"] = "Sigma/Netflex"

        return final_product

    def _process_tci_product(self, tci_product: tci.Product, context: Dict = None) -> Dict[str, Any]:
        parities = self.currency_converter.get_parities()
        tci_coefficient = self.settings.get('tci_coefficient', 1.4)
        processed_variations = []
        all_price_options = []
        valid_tci_variations_exist = False

        if not tci_product.variations:
            logging.warning(f"TCI ürünü ({tci_product.code}) için varyasyon bulunamadı.")
        else:
            for variation in tci_product.variations:
                original_price_str = variation.get('price', 'N/A')
                currency_symbol = '€' if '€' in original_price_str else '$' if '$' in original_price_str else '£' if '£' in original_price_str else '€'
                price_float = None
                calculated_price_eur = None

                try:
                    cleaned = re.sub(r'[^\d,.]', '', original_price_str)
                    if cleaned:
                        standardized = cleaned.replace('.', '').replace(',', '.') if cleaned.rfind(',') > cleaned.rfind(
                            '.') else cleaned.replace(',', '')
                        price_float = float(standardized) if standardized else None

                    if price_float is not None:
                        base_price_eur = price_float
                        if currency_symbol == '$' and parities.get('usd_eur'):
                            base_price_eur = price_float * parities['usd_eur']
                        elif currency_symbol == '£' and parities.get('gbp_eur'):
                            base_price_eur = price_float * parities['gbp_eur']
                        calculated_price_eur = base_price_eur * tci_coefficient
                        valid_tci_variations_exist = True
                        all_price_options.append({
                            'price': calculated_price_eur,
                            'code': f"{tci_product.code}-{variation.get('unit', 'N/A')}",
                            'source': 'TCI'
                        })

                except (ValueError, TypeError) as e:
                    logging.warning(
                        f"TCI fiyat parse/hesaplama hatası ({tci_product.code} - {variation.get('unit')}): {e}")

                processed_variations.append({
                    "unit": variation.get('unit'),
                    "original_price": original_price_str,
                    "original_price_numeric": price_float,
                    "stock_info": variation.get('stock_info', []),
                    "calculated_price_eur": calculated_price_eur,
                    "calculated_price_eur_str": f"{calculated_price_eur:,.2f}€".replace(",", "X").replace(".",
                                                                                                          ",").replace(
                        "X", ".") if calculated_price_eur is not None else "N/A"
                })

        cheapest_option = min(all_price_options, key=lambda x: x['price']) if all_price_options else {}

        final_product = {
            "source": "TCI",
            "product_name": tci_product.name or "N/A",
            "product_number": tci_product.code or "N/A",
            "cas_number": tci_product.cas_number or "N/A",
            "brand": "TCI",
            "tci_variations": processed_variations if valid_tci_variations_exist else [],
            "sigma_variations": {},
            "netflex_matches": [],
            "cheapest_eur_price_str": cheapest_option.get('price') and f"{cheapest_option['price']:,.2f}€".replace(",",
                                                                                                                   "X").replace(
                ".", ",").replace("X", ".") or "Fiyat Yok",
            "cheapest_material_number": cheapest_option.get('code', tci_product.code or "N/A"),
            "cheapest_source_country": cheapest_option.get('source', "TCI")
        }

        cheapest_variation_details = next((v for v in processed_variations if
                                           v.get('calculated_price_eur_str') == final_product[
                                               "cheapest_eur_price_str"]), None)
        if cheapest_variation_details and cheapest_variation_details.get('stock_info'):
            final_product["cheapest_netflex_stock"] = ", ".join(
                [f"{s['country']}: {s['stock']}" for s in cheapest_variation_details['stock_info']])
        else:
            final_product["cheapest_netflex_stock"] = "N/A"

        return final_product

    def _process_orkim_product(self, orkim_product: Dict[str, Any], context: Dict = None) -> Dict[str, Any]:
        """Orkim'den gelen ürün verisini standart formata çevirir."""
        # BU FONKSİYON İÇİNDEKİ STOK MANTIĞI ÖNCEKİ ADIMDA GÜNCELLENDİ (DETAYLI SORGULAMA KALDIRILDI)
        stock_quantity = orkim_product.get("stock_quantity")  # Bu artık "Var" veya 0 olabilir
        stock_status = orkim_product.get("stock_status")
        stock_display = "N/A"
        if stock_status == "Stokta Yok" or stock_quantity == 0:
            stock_display = 0  # 0 olarak gönder
        elif stock_quantity == "Var":  # "Var" metnini işle
            stock_display = "Var"
        elif isinstance(stock_quantity, int):  # _get_stock_from_page çağrılırsa diye (ürün sayfasında)
            stock_display = stock_quantity
        else:
            stock_display = orkim_product.get("stock_quantity", "N/A")  # Orijinal değeri koru

        price_str = orkim_product.get("price_str", "N/A")
        return {
            "source": "Orkim",
            "product_name": orkim_product.get("urun_adi", "N/A"),
            "product_number": orkim_product.get("k_kodu", "N/A"),
            "cas_number": "N/A",
            "brand": orkim_product.get("brand", "Orkim"),
            "cheapest_eur_price_str": price_str,
            "cheapest_material_number": orkim_product.get("k_kodu", "N/A"),
            "cheapest_source_country": "Orkim",
            "cheapest_netflex_stock": stock_display,
            "sigma_variations": {},
            "netflex_matches": [],
            "tci_variations": [],
            "product_url": orkim_product.get("product_url")  # YENİ EKLENEN SATIR
        }

    def _process_itk_product(self, itk_product: Dict[str, Any], context: Dict = None) -> Dict[str, Any]:
        """ITK'den gelen ürün verisini standart formata çevirir ve para birimini EUR'ya dönüştürür."""
        original_price = itk_product.get("price")
        original_currency = itk_product.get("currency", "EUR").upper()
        eur_price = None
        cheapest_price_str = "Fiyat Yok"

        if original_price is not None:
            parities = self.currency_converter.get_parities()
            if "error" in parities: logging.warning("ITK fiyat dönüşümü için pariteler alınamadı.")

            if original_currency == "EUR":
                eur_price = original_price
            elif original_currency == "USD" and parities.get("usd_eur"):
                eur_price = original_price * parities["usd_eur"]
            elif original_currency == "GBP" and parities.get("gbp_eur"):
                eur_price = original_price * parities["gbp_eur"]
            else:
                if original_currency != "EUR": logging.warning(
                    f"ITK için {original_currency} -> EUR dönüşüm oranı bulunamadı.")
                eur_price = original_price

            if eur_price is not None:
                eur_price = eur_price * self.settings.get('itk_coefficient', 1.0)
                cheapest_price_str = f"{eur_price:,.2f}€".replace(",", "X").replace(".", ",").replace("X", ".")

        stock_quantity = itk_product.get("stock_quantity", "N/A")

        itk_variation_data = {
            "product_code": itk_product.get("product_code", "N/A"),
            "product_name": itk_product.get("product_name", "N/A"),
            "price_str": cheapest_price_str,
            "price": eur_price,
            "currency": "EUR",
            "stock_quantity": stock_quantity
        }

        return {
            "source": "ITK",
            "product_name": itk_product.get("product_name", "N/A"),
            "product_number": itk_product.get("product_code", "N/A"),
            "cas_number": "N/A",
            "brand": "ITK",
            "cheapest_eur_price_str": cheapest_price_str,
            "cheapest_material_number": itk_product.get("product_code", "N/A"),
            "cheapest_source_country": "ITK",
            "cheapest_netflex_stock": stock_quantity,
            "sigma_variations": {},
            "netflex_matches": [],
            "tci_variations": [],
            "itk_variations": [itk_variation_data]
        }

    def _process_netflex_product(self, netflex_product: Dict[str, Any], context: Dict = None) -> Dict[str, Any]:
        """Doğrudan Netflex aramasından gelen bir ürünü standart formata çevirir."""
        price_str = netflex_product.get("price_str", "N/A")
        return {
            "source": "Netflex",
            "product_name": netflex_product.get("product_name", "N/A"),
            "product_number": netflex_product.get("product_code", "N/A"),
            "cas_number": "N/A",
            "brand": netflex_product.get("brand", "Netflex"),
            "cheapest_eur_price_str": price_str,
            "cheapest_material_number": netflex_product.get("product_code", "N/A"),
            "cheapest_source_country": "Netflex",
            "cheapest_netflex_stock": netflex_product.get("stock", "N/A"),
            "sigma_variations": {},
            "netflex_matches": [],
            "tci_variations": [],
            "itk_variations": []
        }

    # DEĞİŞİKLİK: 'search_term: str' yerine 'search_data: dict' al
    def search_and_compare(self, search_data: dict, context: Dict = None):
        start_time = time.monotonic()

        # DEĞİŞİKLİK: search_term ve search_logic'i 'search_data' objesinden al
        search_term = search_data.get("searchTerm", "")
        search_logic = search_data.get("searchLogic",
                                       "exact")  # Kullanıcının istediği esnek mantık için "exact" kalıyor

        logging.info(f"ANLIK ARAMA BAŞLATILDI: '{search_term}' (Mantık: {search_logic})")
        if not context: send_to_frontend("log_search_term", {"term": search_term}); admin_logger.info(
            f"Arama: '{search_term}' (Mantık: {search_logic})")

        # M-kodu varyasyonlarını yönetme
        search_term_variations = {search_term.lower()}
        normalized_term = search_term.lower().strip()
        is_m_code = False
        if normalized_term.startswith('m'):
            is_m_code = True
            if '.' in normalized_term:
                search_term_variations.add(normalized_term.replace('.', '', 1))
            elif re.match(r'^m\d+$', normalized_term):
                search_term_variations.add(f"m.{normalized_term[1:]}")

        total_found = 0
        total_found_lock = threading.Lock()
        sigma_found_count = 0
        sigma_found_lock = threading.Lock()

        # 1. Aşama: Sigma, TCI, Orkim ve ITK'da paralel arama
        with ThreadPoolExecutor(max_workers=5, thread_name_prefix="Source-Streamer") as executor:
            def tci_task():
                nonlocal total_found
                try:
                    for product_page in self.tci_api.get_products(search_term, self.search_cancelled):
                        if self.search_cancelled.is_set(): break
                        for product in product_page:
                            if self.search_cancelled.is_set(): break

                            # KULLANICININ İSTEDİĞİ ESNEK "exact" FİLTRELEME (TCI)
                            match_found = False
                            term_lower = search_term.lower()
                            product_name_lower = (product.name or "").lower()
                            product_code_lower = (product.code or "").lower()
                            cas_number_lower = (product.cas_number or "").lower()

                            if search_logic == "exact":
                                # --- KULLANICININ ORİJİNAL MANTIĞI KORUNDU ---
                                if (term_lower == product_name_lower or
                                        (term_lower in product_code_lower or product_code_lower in term_lower) or
                                        (cas_number_lower and term_lower == cas_number_lower)):
                                    match_found = True
                                # --- KULLANICININ ORİJİNAL MANTIĞI SONU ---
                            else:
                                match_found = True

                            if match_found:
                                processed = self._process_tci_product(product, context)
                                send_to_frontend("product_found", {"product": processed}, context=context)
                                with total_found_lock:
                                    total_found += 1
                except Exception as e:
                    logging.error(f"TCI akış hatası: {e}", exc_info=True)

            def sigma_task():
                nonlocal total_found, sigma_found_count
                with ThreadPoolExecutor(max_workers=self.max_workers,
                                        thread_name_prefix="Sigma-Processor") as processor:
                    try:
                        self.currency_converter.get_parities()
                        raw_product_stream = self.sigma_api.search_products(search_term, self.search_cancelled)
                        futures = []
                        for raw_product in raw_product_stream:
                            if self.search_cancelled.is_set(): break
                            # Görevi gönderirken search_data'yı da ekle
                            futures.append(
                                processor.submit(self._process_single_sigma_product_and_send, raw_product, context,
                                                 search_data))
                        for future in as_completed(futures):
                            if future.result():
                                with total_found_lock: total_found += 1
                                with sigma_found_lock: sigma_found_count += 1
                    except Exception as e:
                        logging.error(f"Sigma akış hatası: {e}", exc_info=True)

            def orkim_task():
                nonlocal total_found
                try:
                    # ORKİM OPTİMİZASYONU: Arka plan yöneticisi olduğu için _login çağrısı yok
                    if self.orkim_api:
                        orkim_search_term = search_term
                        # search_logic'i Orkim API'sine gönder
                        orkim_results = self.orkim_api.search_products(orkim_search_term, self.search_cancelled,
                                                                       search_logic)
                        if self.search_cancelled.is_set(): return

                        # Orkim scraper artık filtrelemeyi kendi içinde yapıyor (esnek exact mantığıyla).
                        for product in orkim_results:
                            if self.search_cancelled.is_set(): break
                            processed = self._process_orkim_product(product, context)
                            send_to_frontend("product_found", {"product": processed}, context=context)
                            with total_found_lock:
                                total_found += 1
                except Exception as e:
                    logging.error(f"Orkim akış hatası: {e}", exc_info=True)

            def itk_task():
                nonlocal total_found
                itk_search_terms = {search_term.lower()}
                if is_m_code: itk_search_terms = search_term_variations
                found_codes = set()
                with itk_cache_lock:
                    cache_to_search = list(itk_product_cache)

                term_lower = search_term.lower()

                for product in cache_to_search:
                    if self.search_cancelled.is_set(): return
                    code_lower = product.get("product_code", "").lower()
                    name_lower = product.get("product_name", "").lower()

                    # KULLANICININ İSTEDİĞİ ESNEK "exact" FİLTRELEME (ITK)
                    match_found = False
                    if search_logic == "exact":
                        # --- KULLANICININ ORİJİNAL MANTIĞI KORUNDU ---
                        if (term_lower == name_lower or
                                (term_lower in code_lower or code_lower in term_lower)):
                            match_found = True
                        # --- KULLANICININ ORİJİNAL MANTIĞI SONU ---
                    else:
                        score = 100 if term_lower == code_lower else max(fuzz.partial_ratio(term_lower, name_lower),
                                                                         fuzz.partial_ratio(term_lower, code_lower))
                        if score > 85:
                            match_found = True

                    if match_found and code_lower not in found_codes:
                        processed = self._process_itk_product(product, context)
                        send_to_frontend("product_found", {"product": processed}, context=context)
                        with total_found_lock: total_found += 1
                        found_codes.add(code_lower)

            f_tci = executor.submit(tci_task)
            f_sigma = executor.submit(sigma_task)
            f_orkim = executor.submit(orkim_task)
            f_itk = executor.submit(itk_task)
            # Görevlerin bitmesini bekle
            f_tci.result()
            f_sigma.result()
            f_orkim.result()
            f_itk.result()

        # 2. Aşama: Eğer ilk aşamada HİÇBİR yerden sonuç bulunamadıysa Netflex'i dene
        if total_found == 0 and not self.search_cancelled.is_set():
            logging.info(f"İlk aşamada sonuç bulunamadı, şimdi Netflex'te aranıyor: '{search_term}'")
            try:
                # Dikkat: Netflex API'sinden AuthenticationError gelebilir
                netflex_results = self.netflex_api.search_products(search_term, self.search_cancelled)
                if not self.search_cancelled.is_set():

                    term_lower = search_term.lower()

                    for product in netflex_results:
                        if self.search_cancelled.is_set(): break

                        # KULLANICININ İSTEDİĞİ ESNEK "exact" FİLTRELEME (Netflex 2. Aşama)
                        match_found = False
                        product_name_lower = (product.get('product_name', '') or "").lower()
                        product_code_lower = (product.get('product_code', '') or "").lower()

                        if search_logic == "exact":
                            # --- KULLANICININ ORİJİNAL MANTIĞI KORUNDU ---
                            if (term_lower == product_name_lower or
                                    (term_lower in product_code_lower or product_code_lower in term_lower)):
                                match_found = True
                            # --- KULLANICININ ORİJİNAL MANTIĞI SONU ---
                        else:
                            match_found = True

                        if match_found:
                            processed = self._process_netflex_product(product, context)
                            send_to_frontend("product_found", {"product": processed}, context=context)
                            with total_found_lock:
                                total_found += 1
            except netflex.AuthenticationError:
                logging.error("Netflex kimlik doğrulaması başarısız oldu (2. aşama Netflex araması).")
            except Exception as e:
                logging.error(f"İkincil Netflex araması sırasında hata: {e}", exc_info=True)

        # Arama tamamlama mesajını en sonda gönder
        if not self.search_cancelled.is_set():
            logging.info(
                f"Arama Tamamlandı: '{search_term}', Toplam={total_found}, Süre={time.monotonic() - start_time:.2f}s")
            send_to_frontend("search_complete", {"status": "complete", "total_found": total_found}, context=context)
        elif not context:  # Sadece anlık arama iptal edildiğinde gönder
            send_to_frontend("search_complete", {"status": "cancelled"})
            logging.warning(f"Arama İptal Edildi: '{search_term}'")

    def run_batch_search(self, file_path, customer_name):
        logging.info(f"Toplu arama: Dosya={file_path}, Müşteri={customer_name}")
        self.batch_search_cancelled.clear()
        admin_logger.info(f"Toplu Arama: Müşteri='{customer_name}', Dosya='{os.path.basename(file_path)}'")
        search_terms = get_search_terms_from_file(file_path)
        if not search_terms:
            send_to_frontend("batch_search_complete", {"status": "error", "message": "Dosyadan ürün okunamadı."})
            return
        total_terms = len(search_terms)
        for i, term in enumerate(search_terms):
            if self.batch_search_cancelled.is_set(): logging.warning("Toplu arama iptal edildi."); break
            self.search_cancelled.clear()  # Her terim için anlık arama iptalini temizle
            send_to_frontend("log_search_term", {"term": term})
            send_to_frontend("batch_search_progress", {"term": term, "current": i + 1, "total": total_terms})
            admin_logger.info(f"  -> Toplu Arama ({i + 1}/{total_terms}): '{term}'")

            # Toplu arama her zaman "benzer" mantığını kullanır
            search_data = {"searchTerm": term, "searchLogic": "similar"}
            self.search_and_compare(search_data, context={"batch_search_term": term})
            # İptal edilip edilmediğini kontrol et
            if self.search_cancelled.is_set() and not self.batch_search_cancelled.is_set():
                logging.info(f"'{term}' araması atlandı (cancel_current_term).")
                continue  # Sonraki terime geç

        status = "cancelled" if self.batch_search_cancelled.is_set() else "complete"
        send_to_frontend("batch_search_complete", {"status": status})
        if status == 'complete': admin_logger.info(f"Toplu Arama Tamamlandı: Müşteri='{customer_name}'")

    def force_cancel(self):
        """Anlık aramayı (veya toplu aramadaki mevcut terimi) iptal eder."""
        self.search_cancelled.set()
        logging.info("Anlık arama iptal sinyali gönderildi.")

    def force_cancel_batch(self):
        """Tüm toplu aramayı iptal eder."""
        self.batch_search_cancelled.set()
        self.force_cancel()  # Mevcut çalışan terimi de iptal et
        logging.info("Toplu arama iptal sinyali gönderildi.")


# --- Ana Fonksiyon ve Komut Döngüsü (DÜZELTİLDİ) ---
def main():
    logging.info("=" * 40 + "\nPython Arka Plan Servisi Başlatıldı\n" + "=" * 40)
    start_notification_scheduler()
    services_initialized = threading.Event()
    # Scraper instance'larını başta oluştur
    sigma_api = sigma.SigmaAldrichAPI()
    tci_api = tci.TciScraper()
    currency_api = currency_converter.CurrencyConverter()
    # ITK ve Orkim'i None olarak başlat, ayarlardan sonra oluşturulacak
    itk_api = None
    orkim_api = None
    netflex_api = None
    engine = None
    search_thread = None
    batch_search_thread = None

    def _populate_itk_cache(api_instance):
        """ITK ürünlerini çekip global önbelleği dolduran thread fonksiyonu."""
        if not api_instance: return  # API instance yoksa çık
        logging.info("ITK ürün önbelleği oluşturuluyor...")
        start_time = time.monotonic()
        try:
            products = api_instance.get_all_products()
            with itk_cache_lock:
                global itk_product_cache
                itk_product_cache = products
            duration = time.monotonic() - start_time
            logging.info(f"ITK önbelleği {len(products)} ürünle {duration:.2f} saniyede tamamlandı.")
        except Exception as e:
            logging.error(f"ITK önbelleği oluşturulurken hata: {e}", exc_info=True)

    def initialize_services(settings_data: Dict[str, Any]):
        nonlocal netflex_api, engine, orkim_api, itk_api  # Global değişkenleri kullanacağımızı belirt
        logging.info(f"Servisler başlatılıyor...")
        try:
            # Netflex API'sini oluştur veya güncelle
            if netflex_api:
                netflex_api.update_credentials(settings_data.get("netflex_username"),
                                               settings_data.get("netflex_password"))
            else:
                netflex_api = netflex.NetflexAPI(username=settings_data.get("netflex_username"),
                                                 password=settings_data.get("netflex_password"))

            # Orkim API'sini oluştur veya güncelle
            if orkim_api:
                orkim_api.username = settings_data.get("orkim_username")
                orkim_api.password = settings_data.get("orkim_password")
                orkim_api.openai_api_key = os.getenv("OCR_API_KEY")
                orkim_api.is_logged_in = False  # Tekrar giriş yapmayı zorunlu kıl
            else:
                orkim_api = orkim.OrkimScraper(
                    username=settings_data.get("orkim_username"),
                    password=settings_data.get("orkim_password"),
                    openai_api_key=os.getenv("OCR_API_KEY")
                )
                # Orkim arka plan yöneticisini SADECE İLK OLUŞTURMADA başlat
                threading.Thread(
                    target=orkim_api.run_background_session_manager,
                    name="Orkim-Session-Manager",
                    daemon=True
                ).start()

            # ITK API'sini oluştur veya güncelle
            if itk_api:
                itk_api.USERNAME = settings_data.get("itk_username")
                itk_api.PASSWORD = settings_data.get("itk_password")
            else:
                itk_api = itk.ItkScraper(username=settings_data.get("itk_username"),
                                         password=settings_data.get("itk_password"))
                # ITK önbelleğini SADECE İLK OLUŞTURMADA doldur
                threading.Thread(target=_populate_itk_cache, args=(itk_api,), name="ITK-Cache-Builder",
                                 daemon=True).start()

            # Comparison Engine'i oluştur veya güncelle
            if engine:
                engine.settings = settings_data
                engine.netflex_api = netflex_api
                engine.orkim_api = orkim_api
                engine.itk_api = itk_api
            else:
                engine = ComparisonEngine(sigma_api, netflex_api, tci_api, orkim_api, itk_api,
                                          initial_settings=settings_data)

            # Servis başlatma görevini (ağır olanlar) ayrı thread'de çalıştır
            def init_task():
                try:
                    # Netflex token almayı dene (kimlik doğrulaması kontrolü)
                    netflex_api.get_token()
                    # Sigma ve TCI sürücülerini başlat (eğer engine varsa ve daha önce başlatılmadıysa)
                    # Not: Bu, her ayar kaydında sürücüleri yeniden başlatmaz, sadece ilk seferde yapar.
                    # Eğer sürücülerin yeniden başlatılması gerekiyorsa farklı bir mantık gerekir.
                    if engine and not services_initialized.is_set():  # Sadece ilk başlatmada veya hatadan sonra
                        engine.initialize_drivers()

                    send_to_frontend("python_services_ready", True)
                    services_initialized.set()  # Servisler hazır
                except netflex.AuthenticationError:
                    send_to_frontend("authentication_error", True)
                    services_initialized.clear()  # Hazır değil
                    send_to_frontend("python_services_ready", False)
                # except (wmi.x_wmi, requests.RequestException) as license_e: # Lisans kontrolü hatası
                #      logging.error(f"Lisans kontrolü başarısız: {license_e}")
                #      send_to_frontend("license_error", True)
                #      services_initialized.clear()
                #      send_to_frontend("python_services_ready", False)
                except Exception as e:  # Diğer başlatma hataları (örn: Selenium)
                    logging.critical(f"Servis başlatma alt görevi (init_task) hatası: {e}", exc_info=True)
                    send_to_frontend("python_services_ready", False)
                    services_initialized.clear()  # Hazır değil

            threading.Thread(target=init_task, name="Full-Initializer", daemon=True).start()
        except Exception as e:
            logging.critical(f"Ana servis başlatma (initialize_services) hatası: {e}", exc_info=True)
            send_to_frontend("python_services_ready", False)  # Ana başlatmada hata olursa da hazır değil gönder

    # İlk ayarları yükle ve servisleri başlatmayı dene
    loaded_settings, _ = load_settings()
    if not all([loaded_settings.get(k) for k in
                ["netflex_username", "netflex_password", "orkim_username", "orkim_password", "itk_username",
                 "itk_password"]]):  # Temel bilgiler eksikse
        send_to_frontend("initial_setup_required", True)
    else:
        initialize_services(loaded_settings)

    # stdin'den komutları dinle
    for line_bytes in sys.stdin.buffer:
        line = line_bytes.decode('utf-8', errors='replace')
        if not line.strip():
            continue
        try:
            request = json.loads(line.strip())
            action, data = request.get("action"), request.get("data")
            logging.debug(f"Komut alındı: Eylem='{action}'")  # INFO yerine DEBUG

            if action == "load_settings":
                settings_data, was_upgraded = load_settings()
                send_to_frontend("settings_loaded", settings_data)
                if was_upgraded:
                    send_to_frontend("new_settings_available", True)
            elif action == "save_settings" and isinstance(data, dict):
                save_settings(data)
                # Ayarlar kaydedildiğinde servisleri yeniden başlat VEYA güncelle
                logging.info("Ayarlar kaydedildi, servisler güncelleniyor/yeniden başlatılıyor...")
                services_initialized.clear()  # Yeniden başlatma için flag'i temizle
                initialize_services(data)  # Bu fonksiyon artık mevcut instance'ları güncelliyor
                send_to_frontend("settings_saved", {"status": "success"})

            elif action == "load_calendar_notes":
                send_to_frontend("calendar_notes_loaded", load_calendar_notes())
            elif action == "save_calendar_notes" and isinstance(data, list):
                save_calendar_notes(data)
                # Bildirimleri hemen kontrol etmeye gerek yok, periyodik kontrol yeterli
                # threading.Thread(target=_perform_notification_check, name="Manual-Notification-Check").start()
                send_to_frontend("calendar_notes_saved", {"status": "success"})
            elif action == "mark_meeting_complete" and isinstance(data, dict):
                if data.get("noteDate") and data.get("meetingId"):
                    _mark_meeting_as_complete(data["noteDate"], data["meetingId"])
            elif action in ["search", "start_batch_search"] and data:
                if not services_initialized.is_set():
                    # Eğer servisler hazır değilse, hata mesajı gönder
                    if not netflex_api or not netflex_api.credentials.get(
                            "adi"):  # Temel bilgiler eksikse kurulum gerekli
                        send_to_frontend("initial_setup_required", True)
                    else:  # Başlatma sırasında bir hata oluşmuş olabilir
                        send_to_frontend("search_error",
                                         "Servisler henüz başlatılmadı veya başlatılırken hata oluştu. Lütfen ayarları kontrol edin veya uygulamayı yeniden başlatın.")
                    continue

                # Mevcut aramaları durdur (varsa)
                if engine:
                    if search_thread and search_thread.is_alive():
                        logging.debug("Önceki anlık arama durduruluyor...")
                        engine.force_cancel()
                        search_thread.join(2.0)  # Kısa bir süre bekle
                    if batch_search_thread and batch_search_thread.is_alive():
                        logging.debug("Önceki toplu arama durduruluyor...")
                        engine.force_cancel_batch()
                        batch_search_thread.join(2.0)  # Kısa bir süre bekle

                    # Yeni aramayı başlat
                    if action == "search":
                        engine.search_cancelled.clear()
                        search_thread = threading.Thread(target=engine.search_and_compare, args=(data,),
                                                         name="Search-Coordinator", daemon=True)
                        search_thread.start()
                    else:  # start_batch_search
                        engine.batch_search_cancelled.clear()
                        engine.search_cancelled.clear()  # Her ihtimale karşı anlık iptali de temizle
                        batch_search_thread = threading.Thread(target=engine.run_batch_search,
                                                               args=(data.get("filePath"), data.get("customerName")),
                                                               name="Batch-Search-Coordinator", daemon=True)
                        batch_search_thread.start()
                else:  # Engine başlatılamadıysa
                    send_to_frontend("search_error", "Arama motoru başlatılamadı. Ayarları kontrol edin.")


            elif action == "cancel_search" or action == "cancel_current_term_search":
                if engine: engine.force_cancel()
            elif action == "cancel_batch_search":
                if engine: engine.force_cancel_batch()
            elif action == "export":
                send_to_frontend("export_result", export_to_excel(data))
            elif action == "export_meetings":
                send_to_frontend("export_meetings_result", export_meetings_to_excel(data))
            elif action == "get_parities":
                send_to_frontend("parities_updated", currency_api.get_parities())

            # YENİ: Orkim Stok Sorgulama
            elif action == "get_orkim_stock" and isinstance(data, dict) and data.get("url"):
                if not orkim_api:
                    logging.warning("Orkim API hazır değilken stok sorgusu istendi.")
                    send_to_frontend("orkim_stock_result", {"url": data.get("url"), "stock": "Hata"})
                else:
                    # Yavaş işlemi ana döngüyü kilitlememesi için bir thread'de başlat
                    threading.Thread(
                        target=_get_orkim_stock_task,
                        args=(orkim_api, data.get("url")),
                        name="Orkim-Stock-Check",
                        daemon=True
                    ).start()

            elif action == "shutdown":
                logging.info("Kapatma komutu alındı. Kaynaklar serbest bırakılıyor...")
                stop_notification_scheduler()

                # Önce çalışan arama thread'lerini durdurmaya çalış
                if engine:
                    engine.force_cancel()
                    engine.force_cancel_batch()
                    if search_thread and search_thread.is_alive(): search_thread.join(1.0)
                    if batch_search_thread and batch_search_thread.is_alive(): batch_search_thread.join(1.0)

                # Sonra Selenium sürücülerini kapat
                driver_shutdown_errors = False
                try:
                    if sigma_api: sigma_api.stop_drivers()
                except Exception as e:
                    logging.error(f"Sigma sürücüleri kapatılırken hata: {e}"); driver_shutdown_errors = True
                try:
                    if tci_api: tci_api.close_driver()
                except Exception as e:
                    logging.error(f"TCI sürücüsü kapatılırken hata: {e}"); driver_shutdown_errors = True
                try:
                    if orkim_api: orkim_api.close_driver()  # Session'ı kapatır ve arka plan thread'ine sinyal gönderir
                except Exception as e:
                    logging.error(f"Orkim oturumu kapatılırken hata: {e}"); driver_shutdown_errors = True

                if driver_shutdown_errors:
                    logging.warning("Bazı sürücüler kapatılırken hata oluştu.")
                else:
                    logging.info("Tüm sürücüler ve oturumlar başarıyla kapatıldı.")

                logging.info("Arka plan servisinden çıkılıyor.")
                send_to_frontend('python_shutdown_complete', {})
                sys.stdout.flush()
                time.sleep(0.1)
                break  # Ana döngüden çık ve script'i sonlandır

        except json.JSONDecodeError:
            logging.error(f"Geçersiz JSON alındı: {line.strip()}")
        except Exception as e:
            logging.critical(f"Ana döngüde beklenmedik bir hata oluştu: {e}", exc_info=True)

    # Döngü bittiğinde (shutdown komutuyla) veya bir hata oluştuğunda buraya gelinir.
    logging.info("Python ana döngüsü sona erdi.")
    stop_notification_scheduler()


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        logging.critical(f"main() fonksiyonunda yakalanmayan kritik hata: {e}", exc_info=True)
        # Kritik hatada bile Electron'a kapanması gerektiğini söyle
        send_to_frontend('python_shutdown_complete', {'error': True})
        sys.stdout.flush()
        time.sleep(0.1)
